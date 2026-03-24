"""Tests for the Filtron Excel catalog parser.

Every test creates synthetic .xlsx workbooks via openpyxl so the suite is
self-contained (no fixture files needed).
"""

import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook

from parts_finder.models import ProductCrossRef
from parts_finder.parsers.filtron_parser import (
    OPEN_ENDED_YEAR,
    FiltronParser,
    FiltronRecord,
    parse_year_range,
    resolve_filter_type,
    split_multi_part,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_workbook(headers: list[str], rows: list[list], *, title_rows: int = 0) -> Path:
    """Create a temporary .xlsx file and return its path.

    ``title_rows`` inserts that many blank/title rows above the header so
    we can test header auto-detection on non-first rows.
    """
    wb = Workbook()
    ws = wb.active
    for _ in range(title_rows):
        ws.append(["Filtron Catalog 2024"])
    ws.append(headers)
    for row in rows:
        ws.append(row)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    wb.close()
    tmp.close()
    return Path(tmp.name)


# A reusable "standard" workbook for basic tests
_STANDARD_HEADERS = ["Make", "Model", "Year", "Engine", "Filter Type", "Filtron", "OEM No.", "MANN-FILTER"]
_STANDARD_ROWS = [
    ["Toyota", "Corolla", "2019-2023", "2ZR-FE", "Oil", "OP 621", "04152-YZZA1", "W 712/95"],
    ["Toyota", "Corolla", "2019-2023", "2ZR-FE", "Air", "AP 177/7", "17801-21050", "C 1858"],
    ["VW", "Golf", "2015-", "CZCA", "Oil", "OP 641/2", "04E115561H, 04E115561T", "W 712/94"],
]


# ===========================================================================
# Test classes
# ===========================================================================


class TestParseBasicExcel:
    """Parse a synthetic file, verify record count and field values."""

    def test_record_count(self, tmp_path):
        path = _make_workbook(_STANDARD_HEADERS, _STANDARD_ROWS)
        records = FiltronParser(path).parse()
        assert len(records) == 3

    def test_first_record_fields(self, tmp_path):
        path = _make_workbook(_STANDARD_HEADERS, _STANDARD_ROWS)
        rec = FiltronParser(path).parse()[0]
        assert rec.make == "Toyota"
        assert rec.model == "Corolla"
        assert rec.year_from == 2019
        assert rec.year_to == 2023
        assert rec.engine_code == "2ZR-FE"
        assert rec.filter_type == "oil_filter"
        assert rec.filtron_part == "OP 621"
        assert rec.oem_parts == ["04152-YZZA1"]
        assert rec.mann_part == "W 712/95"

    def test_vw_open_ended_year(self, tmp_path):
        path = _make_workbook(_STANDARD_HEADERS, _STANDARD_ROWS)
        rec = FiltronParser(path).parse()[2]
        assert rec.make == "VW"
        assert rec.year_to == OPEN_ENDED_YEAR

    def test_vw_multi_oem(self, tmp_path):
        path = _make_workbook(_STANDARD_HEADERS, _STANDARD_ROWS)
        rec = FiltronParser(path).parse()[2]
        assert rec.oem_parts == ["04E115561H", "04E115561T"]


class TestYearRangeParsing:
    """7 subtests: dash, open-ended, 'from', single year, spaces, en-dash, invalid."""

    def test_dash_range(self):
        assert parse_year_range("2019-2023") == (2019, 2023)

    def test_open_ended(self):
        assert parse_year_range("2019-") == (2019, OPEN_ENDED_YEAR)

    def test_from_keyword(self):
        assert parse_year_range("from 2019") == (2019, OPEN_ENDED_YEAR)

    def test_single_year(self):
        assert parse_year_range("2021") == (2021, 2021)

    def test_spaces(self):
        assert parse_year_range("2019 - 2023") == (2019, 2023)

    def test_en_dash(self):
        assert parse_year_range("2019\u20132023") == (2019, 2023)

    def test_invalid(self):
        assert parse_year_range("N/A") == (0, 0)

    def test_empty(self):
        assert parse_year_range("") == (0, 0)


class TestOemPartExtraction:
    """Whitespace stripping, empty string handling."""

    def test_strips_whitespace(self):
        assert split_multi_part("  ABC123 , DEF456  ") == ["ABC123", "DEF456"]

    def test_empty_string(self):
        assert split_multi_part("") == []

    def test_none_coerced(self):
        assert split_multi_part(None) == []

    def test_semicolon_split(self):
        assert split_multi_part("ABC; DEF;GHI") == ["ABC", "DEF", "GHI"]

    def test_no_slash_split(self):
        """Slashes are NOT delimiters — they appear in part numbers."""
        assert split_multi_part("OX 388D / OX 381D") == ["OX 388D / OX 381D"]


class TestFilterTypeMapping:
    """Oil/air/cabin/fuel variants (English + Polish)."""

    def test_english_oil(self):
        assert resolve_filter_type("Oil", "") == "oil_filter"

    def test_english_air_filter(self):
        assert resolve_filter_type("Air Filter", "") == "air_filter"

    def test_polish_cabin(self):
        assert resolve_filter_type("filtr kabinowy", "") == "cabin_filter"

    def test_polish_fuel(self):
        assert resolve_filter_type("filtr paliwa", "") == "fuel_filter"

    def test_single_letter_code(self):
        assert resolve_filter_type("O", "") == "oil_filter"
        assert resolve_filter_type("A", "") == "air_filter"
        assert resolve_filter_type("K", "") == "cabin_filter"
        assert resolve_filter_type("F", "") == "fuel_filter"

    def test_unknown_falls_through(self):
        assert resolve_filter_type("something weird", "") == "unknown"


class TestToCrossrefs:
    """Filtron+Mann dual entries, Mann-absent case, multi-OEM, frozen instances."""

    def test_dual_brand_entries(self):
        rec = FiltronRecord(
            make="Toyota", model="Corolla", year_from=2019, year_to=2023,
            engine_code="2ZR-FE", filter_type="oil_filter",
            filtron_part="OP 621", oem_parts=["04152-YZZA1"],
            mann_part="W 712/95",
        )
        refs = FiltronParser.to_crossrefs([rec])
        assert len(refs) == 2
        assert refs[0].brand == "Filtron"
        assert refs[0].brand_part_number == "OP 621"
        assert refs[1].brand == "MANN-FILTER"
        assert refs[1].brand_part_number == "W 712/95"

    def test_no_mann(self):
        rec = FiltronRecord(
            filter_type="oil_filter", filtron_part="OP 621",
            oem_parts=["04152-YZZA1"], mann_part="",
        )
        refs = FiltronParser.to_crossrefs([rec])
        assert len(refs) == 1
        assert refs[0].brand == "Filtron"

    def test_multi_oem_multiplication(self):
        rec = FiltronRecord(
            filter_type="oil_filter", filtron_part="OP 641/2",
            oem_parts=["04E115561H", "04E115561T"], mann_part="W 712/94",
        )
        refs = FiltronParser.to_crossrefs([rec])
        # 2 OEM parts * 2 brands = 4 entries
        assert len(refs) == 4
        oem_nums = [r.oem_part_number for r in refs]
        assert oem_nums.count("04E115561H") == 2
        assert oem_nums.count("04E115561T") == 2

    def test_frozen_instances(self):
        rec = FiltronRecord(
            filter_type="oil_filter", filtron_part="OP 621",
            oem_parts=["04152-YZZA1"],
        )
        refs = FiltronParser.to_crossrefs([rec])
        assert isinstance(refs[0], ProductCrossRef)
        with pytest.raises(AttributeError):
            refs[0].brand = "Other"  # type: ignore[misc]

    def test_unknown_type_skipped(self):
        rec = FiltronRecord(
            filter_type="unknown", filtron_part="XX 999",
            oem_parts=["FAKE123"],
        )
        assert FiltronParser.to_crossrefs([rec]) == []

    def test_empty_filtron_part_skipped(self):
        rec = FiltronRecord(
            filter_type="oil_filter", filtron_part="",
            oem_parts=["FAKE123"],
        )
        assert FiltronParser.to_crossrefs([rec]) == []


class TestEmptyAndMalformedRows:
    """Null rows, whitespace-only rows, bad year still parses."""

    def test_empty_rows_skipped(self):
        headers = ["Make", "Model", "Filtron"]
        rows = [
            ["Toyota", "Corolla", "OP 621"],
            [None, None, None],
            ["", "", ""],
            ["VW", "Golf", "OP 641/2"],
        ]
        path = _make_workbook(headers, rows)
        records = FiltronParser(path).parse()
        assert len(records) == 2

    def test_bad_year_still_parses(self):
        headers = ["Make", "Model", "Year", "Filtron"]
        rows = [["Toyota", "Corolla", "N/A", "OP 621"]]
        path = _make_workbook(headers, rows)
        records = FiltronParser(path).parse()
        assert len(records) == 1
        assert records[0].year_from == 0
        assert records[0].year_to == 0

    def test_whitespace_only_row_skipped(self):
        headers = ["Make", "Model", "Filtron"]
        rows = [
            ["Toyota", "Corolla", "OP 621"],
            ["  ", "  ", "  "],
        ]
        path = _make_workbook(headers, rows)
        records = FiltronParser(path).parse()
        # Whitespace-only cells are stripped to "", so the row becomes
        # effectively empty except for filter_type which becomes "unknown"
        # from resolve_filter_type("", " ").  The row IS kept because
        # make=" " strips to "" but _cell_str makes it "".
        # However any("") on all fields = False, so it IS skipped.
        # Let's just verify count: the Toyota row + the whitespace row.
        # Actually _cell_str strips, so raw values are all "", any() = False -> skipped.
        assert len(records) == 1


class TestHeaderAutoDetection:
    """Header on row 3 (after title rows), no header raises ValueError."""

    def test_header_on_row_3(self):
        headers = ["Make", "Model", "Filtron", "OEM"]
        rows = [["Toyota", "Corolla", "OP 621", "04152-YZZA1"]]
        path = _make_workbook(headers, rows, title_rows=2)
        records = FiltronParser(path).parse()
        assert len(records) == 1
        assert records[0].make == "Toyota"

    def test_no_header_raises(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["random", "columns", "here"])
        ws.append(["foo", "bar", "baz"])

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        wb.close()
        tmp.close()

        with pytest.raises(ValueError, match="No header row found"):
            FiltronParser(Path(tmp.name)).parse()

    def test_insufficient_columns_not_header(self):
        """A row with only 2 recognized columns (< 3) is not a header."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Make", "Model"])  # only 2 — not enough
        ws.append(["Toyota", "Corolla"])

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        wb.close()
        tmp.close()

        with pytest.raises(ValueError, match="No header row found"):
            FiltronParser(Path(tmp.name)).parse()


class TestMultiPartNumberCells:
    """Comma-separated OEM parts in Excel context."""

    def test_comma_separated_in_workbook(self):
        headers = ["Make", "Model", "Filtron", "OEM No."]
        rows = [["VW", "Golf", "OP 641/2", "04E115561H, 04E115561T"]]
        path = _make_workbook(headers, rows)
        rec = FiltronParser(path).parse()[0]
        assert rec.oem_parts == ["04E115561H", "04E115561T"]

    def test_semicolon_separated_in_workbook(self):
        headers = ["Make", "Model", "Filtron", "OEM No."]
        rows = [["BMW", "3 Series", "OP 526/5", "11 42 7 953 129; 11 42 7 566 327"]]
        path = _make_workbook(headers, rows)
        rec = FiltronParser(path).parse()[0]
        assert rec.oem_parts == ["11 42 7 953 129", "11 42 7 566 327"]


class TestToSpecsUpdates:
    """Grouping by vehicle identity, unknown filter type skipped."""

    def test_basic_grouping(self):
        records = [
            FiltronRecord(
                make="Toyota", model="Corolla", year_from=2019, year_to=2023,
                engine_code="2ZR-FE", filter_type="oil_filter",
                filtron_part="OP 621", oem_parts=["04152-YZZA1"],
            ),
            FiltronRecord(
                make="Toyota", model="Corolla", year_from=2019, year_to=2023,
                engine_code="2ZR-FE", filter_type="air_filter",
                filtron_part="AP 177/7", oem_parts=["17801-21050"],
            ),
        ]
        updates = FiltronParser.to_specs_updates(records)
        assert len(updates) == 1
        u = updates[0]
        assert u["make"] == "Toyota"
        assert u["oil_filter_oem"] == "04152-YZZA1"
        assert u["air_filter_oem"] == "17801-21050"

    def test_unknown_type_skipped(self):
        records = [
            FiltronRecord(
                make="Toyota", model="Corolla", year_from=2019, year_to=2023,
                engine_code="2ZR-FE", filter_type="unknown",
                filtron_part="XX 999", oem_parts=["FAKE123"],
            ),
        ]
        updates = FiltronParser.to_specs_updates(records)
        assert len(updates) == 0

    def test_multi_oem_joined(self):
        records = [
            FiltronRecord(
                make="VW", model="Golf", year_from=2015, year_to=OPEN_ENDED_YEAR,
                engine_code="CZCA", filter_type="oil_filter",
                filtron_part="OP 641/2", oem_parts=["04E115561H", "04E115561T"],
            ),
        ]
        updates = FiltronParser.to_specs_updates(records)
        assert len(updates) == 1
        # Sorted alphabetically, joined with comma
        assert updates[0]["oil_filter_oem"] == "04E115561H, 04E115561T"

    def test_different_vehicles_not_merged(self):
        records = [
            FiltronRecord(
                make="Toyota", model="Corolla", year_from=2019, year_to=2023,
                engine_code="2ZR-FE", filter_type="oil_filter",
                filtron_part="OP 621", oem_parts=["04152-YZZA1"],
            ),
            FiltronRecord(
                make="VW", model="Golf", year_from=2015, year_to=OPEN_ENDED_YEAR,
                engine_code="CZCA", filter_type="oil_filter",
                filtron_part="OP 641/2", oem_parts=["04E115561H"],
            ),
        ]
        updates = FiltronParser.to_specs_updates(records)
        assert len(updates) == 2


class TestFilterTypeFromPartPrefix:
    """Infer type from Filtron part number when type column is empty."""

    def test_op_prefix_oil(self):
        assert resolve_filter_type("", "OP 621") == "oil_filter"

    def test_ap_prefix_air(self):
        assert resolve_filter_type("", "AP 177/7") == "air_filter"

    def test_ak_prefix_air(self):
        assert resolve_filter_type("", "AK 218") == "air_filter"

    def test_k_prefix_cabin(self):
        assert resolve_filter_type("", "K 1223A") == "cabin_filter"

    def test_pp_prefix_fuel(self):
        assert resolve_filter_type("", "PP 838/1") == "fuel_filter"

    def test_unknown_two_char_k_prefix_not_cabin(self):
        """KN, KB, etc. should NOT fall through to single-char 'K' match."""
        assert resolve_filter_type("", "KN 555") == "unknown"

    def test_no_prefix_match(self):
        assert resolve_filter_type("", "XY 999") == "unknown"

    def test_empty_part_number(self):
        assert resolve_filter_type("", "") == "unknown"

    def test_in_parsed_workbook(self):
        """When the filter_type column is blank, prefix inference kicks in."""
        headers = ["Make", "Model", "Filter Type", "Filtron", "OEM"]
        rows = [["Toyota", "Corolla", "", "OP 621", "04152-YZZA1"]]
        path = _make_workbook(headers, rows)
        rec = FiltronParser(path).parse()[0]
        assert rec.filter_type == "oil_filter"
