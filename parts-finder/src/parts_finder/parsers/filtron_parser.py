"""Parser for Filtron Excel filter catalogs.

Transforms vendor-specific Excel format into canonical domain objects
(VehicleSpecs field updates and ProductCrossRef instances).

Filtron (filtron.eu) publishes downloadable Excel catalogs containing
cross-references between vehicle applications, OEM filter part numbers,
and aftermarket equivalents (Filtron + sometimes MANN-FILTER).

The parser is deliberately flexible about column naming (English/Polish
variants) and robust to dirty data (empty rows, missing years, merged
title rows above the real header).
"""

from __future__ import annotations

import datetime
import logging
import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from parts_finder.models import ProductCrossRef

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

OPEN_ENDED_YEAR: int = datetime.date.today().year

# Maps every known column header variant (lowercased) to a canonical field.
# English + Polish variants cover the two most common Filtron catalog formats.
COLUMN_ALIASES: dict[str, str] = {
    # make
    "make": "make",
    "marka": "make",
    "brand": "make",
    "manufacturer": "make",
    "producent": "make",
    "car make": "make",
    # model
    "model": "model",
    "car model": "model",
    "typ": "model",
    "type": "model",
    # year range (single column containing "YYYY-YYYY")
    "year": "year_range",
    "years": "year_range",
    "year range": "year_range",
    "rok": "year_range",
    "lata": "year_range",
    "rocznik": "year_range",
    "production years": "year_range",
    "rok produkcji": "year_range",
    # year_from / year_to (separate columns)
    "year from": "year_from",
    "od roku": "year_from",
    "from": "year_from",
    "year to": "year_to",
    "do roku": "year_to",
    "to": "year_to",
    # engine code
    "engine": "engine_code",
    "engine code": "engine_code",
    "silnik": "engine_code",
    "kod silnika": "engine_code",
    "engine type": "engine_code",
    # filter type
    "filter type": "filter_type",
    "type of filter": "filter_type",
    "rodzaj filtra": "filter_type",
    "typ filtra": "filter_type",
    "category": "filter_type",
    "kategoria": "filter_type",
    # filtron part number
    "filtron": "filtron_part",
    "filtron no": "filtron_part",
    "filtron no.": "filtron_part",
    "filtron number": "filtron_part",
    "nr filtron": "filtron_part",
    "numer filtron": "filtron_part",
    "filtron part": "filtron_part",
    # OEM part number(s)
    "oem": "oem_part",
    "oem no": "oem_part",
    "oem no.": "oem_part",
    "oem number": "oem_part",
    "oem part": "oem_part",
    "nr oem": "oem_part",
    "numer oem": "oem_part",
    "original": "oem_part",
    "oe": "oem_part",
    "oe number": "oem_part",
    # MANN-FILTER cross-reference
    "mann": "mann_part",
    "mann-filter": "mann_part",
    "mann filter": "mann_part",
    "mann no": "mann_part",
    "mann no.": "mann_part",
    "nr mann": "mann_part",
}

# Maps raw filter-type strings (lowercased) to canonical categories
# matching VehicleSpecs field naming: oil_filter, air_filter, etc.
FILTER_TYPE_MAP: dict[str, str] = {
    # Oil
    "oil": "oil_filter",
    "oil filter": "oil_filter",
    "filtr oleju": "oil_filter",
    "olej": "oil_filter",
    "engine oil": "oil_filter",
    "o": "oil_filter",
    # Air
    "air": "air_filter",
    "air filter": "air_filter",
    "filtr powietrza": "air_filter",
    "powietrze": "air_filter",
    "engine air": "air_filter",
    "a": "air_filter",
    # Cabin / interior
    "cabin": "cabin_filter",
    "cabin filter": "cabin_filter",
    "interior": "cabin_filter",
    "pollen": "cabin_filter",
    "filtr kabinowy": "cabin_filter",
    "kabinowy": "cabin_filter",
    "k": "cabin_filter",
    # Fuel
    "fuel": "fuel_filter",
    "fuel filter": "fuel_filter",
    "filtr paliwa": "fuel_filter",
    "paliwo": "fuel_filter",
    "f": "fuel_filter",
}

# Infer filter type from Filtron part-number prefix when the type column
# is empty.  Filtron uses a systematic prefix scheme:
#   OP  = oil,  AP/AK = air,  K = cabin,  PP = fuel
_PART_PREFIX_TO_TYPE: dict[str, str] = {
    "OP": "oil_filter",
    "OE": "oil_filter",
    "AP": "air_filter",
    "AK": "air_filter",
    "AM": "air_filter",
    "AR": "air_filter",
    "K": "cabin_filter",
    "PP": "fuel_filter",
    "PM": "fuel_filter",
    "PE": "fuel_filter",
    "PS": "fuel_filter",
}

# Canonical field -> VehicleSpecs OEM field name
_FILTER_CATEGORY_TO_SPECS_FIELD: dict[str, str] = {
    "oil_filter": "oil_filter_oem",
    "air_filter": "air_filter_oem",
    "cabin_filter": "cabin_filter_oem",
    "fuel_filter": "fuel_filter_oem",
}


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------


@dataclass
class FiltronRecord:
    """Intermediate representation of one row in a Filtron catalog.

    Non-frozen so helpers can mutate fields during parsing before the
    record is finalized.
    """

    make: str = ""
    model: str = ""
    year_from: int = 0
    year_to: int = 0
    engine_code: str = ""
    filter_type: str = ""
    filtron_part: str = ""
    oem_parts: list[str] = field(default_factory=list)
    mann_part: str = ""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

# Matches "2019-2023", "2019 - 2023", "2019\u20132023" (en-dash), "2019-",
# "from 2019", bare "2019".
_YEAR_RANGE_RE = re.compile(
    r"(?:from\s+)?(\d{4})\s*[-\u2013\u2014]?\s*(\d{4})?"
)


def parse_year_range(raw: str) -> tuple[int, int]:
    """Extract (year_from, year_to) from a free-form year string.

    Returns (0, 0) when the input is unparseable.
    """
    if not raw:
        return 0, 0
    text = str(raw).strip()
    m = _YEAR_RANGE_RE.search(text)
    if not m:
        return 0, 0
    y_from = int(m.group(1))
    if m.group(2):
        y_to = int(m.group(2))
    elif "-" in text or "\u2013" in text or "\u2014" in text:
        # Open-ended: "2019-"
        y_to = OPEN_ENDED_YEAR
    elif text.lower().startswith("from"):
        # "from 2019" = open-ended (introduced that year, still current)
        y_to = OPEN_ENDED_YEAR
    else:
        # Single year: "2019" -> treat as single-year range
        y_to = y_from
    return y_from, y_to


def resolve_filter_type(raw_type: str, filtron_part: str) -> str:
    """Resolve canonical filter type from the type cell or part prefix.

    Strategy:
    1. Direct lookup in FILTER_TYPE_MAP
    2. Infer from Filtron part-number prefix
    3. Fallback: "unknown"
    """
    if raw_type:
        normalized = raw_type.strip().lower()
        if normalized in FILTER_TYPE_MAP:
            return FILTER_TYPE_MAP[normalized]

    # Infer from part prefix — try longest prefix first (2 chars, then 1)
    if filtron_part:
        part_upper = filtron_part.strip().upper()
        for length in (2, 1):
            prefix = part_upper[:length]
            if prefix not in _PART_PREFIX_TO_TYPE:
                continue
            # Single-char prefix: only match if followed by space or digit
            # (guards against 'KN', 'KB', etc. falling through to 'K')
            if length == 1 and len(part_upper) > 1 and part_upper[1] not in (" ", "0", "1", "2", "3", "4", "5", "6", "7", "8", "9"):
                continue
            return _PART_PREFIX_TO_TYPE[prefix]

    return "unknown"


def split_multi_part(cell_value: str) -> list[str]:
    """Split a cell that may contain multiple OEM part numbers.

    Splits on comma and semicolon (NOT slash — part numbers like
    06J115403Q contain no slashes, but cross-ref codes like
    OX 388D / OX 381D do use slashes as part of the number).
    """
    if not cell_value:
        return []
    parts = re.split(r"[,;]", str(cell_value))
    return [p.strip() for p in parts if p.strip()]


def _cell_str(cell_value: object) -> str:
    """Coerce an openpyxl cell value to a stripped string."""
    if cell_value is None:
        return ""
    return str(cell_value).strip()


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------


class FiltronParser:
    """Parses a Filtron Excel catalog into ``FiltronRecord`` objects.

    Usage::

        parser = FiltronParser(Path("filtron_catalog.xlsx"))
        records = parser.parse()
        crossrefs = FiltronParser.to_crossrefs(records)
        specs_updates = FiltronParser.to_specs_updates(records)
    """

    def __init__(
        self,
        path: Path,
        *,
        sheet_name: str | None = None,
        max_header_scan_rows: int = 10,
    ) -> None:
        self.path = Path(path)
        self.sheet_name = sheet_name
        self.max_header_scan_rows = max_header_scan_rows

    # -- public API --------------------------------------------------------

    def parse(self) -> list[FiltronRecord]:
        """Read the workbook and return parsed records."""
        wb = load_workbook(self.path, read_only=True, data_only=True)
        try:
            ws = wb[self.sheet_name] if self.sheet_name else wb.active
            col_map, header_row = self._detect_header(ws)
            records = self._parse_rows(ws, col_map, header_row)
        finally:
            wb.close()
        logger.info(
            "Parsed %d records from %s", len(records), self.path.name
        )
        return records

    # -- conversion helpers ------------------------------------------------

    @staticmethod
    def to_crossrefs(records: list[FiltronRecord]) -> list[ProductCrossRef]:
        """Convert records into ``ProductCrossRef`` instances.

        Each OEM part number generates a separate ``ProductCrossRef`` per
        aftermarket brand (Filtron, and optionally MANN).
        """
        results: list[ProductCrossRef] = []
        for rec in records:
            if rec.filter_type == "unknown" or not rec.filtron_part:
                continue
            for oem in rec.oem_parts:
                if not oem:
                    continue
                results.append(
                    ProductCrossRef(
                        oem_part_number=oem,
                        category=rec.filter_type,
                        brand="Filtron",
                        brand_part_number=rec.filtron_part,
                    )
                )
                if rec.mann_part:
                    results.append(
                        ProductCrossRef(
                            oem_part_number=oem,
                            category=rec.filter_type,
                            brand="MANN-FILTER",
                            brand_part_number=rec.mann_part,
                        )
                    )
        return results

    @staticmethod
    def to_specs_updates(records: list[FiltronRecord]) -> list[dict]:
        """Group records by vehicle identity and collect OEM parts.

        Returns dicts (not ``VehicleSpecs``) because Filtron catalogs lack
        ``fuel_type``.  The calling import orchestrator decides how to
        merge these partial updates into full ``VehicleSpecs`` rows.

        Records with ``filter_type == "unknown"`` are skipped.
        """
        # Key: (make, model, year_from, year_to, engine_code)
        groups: dict[tuple, dict[str, set[str]]] = defaultdict(
            lambda: defaultdict(set)
        )

        for rec in records:
            if rec.filter_type == "unknown":
                continue
            specs_field = _FILTER_CATEGORY_TO_SPECS_FIELD.get(rec.filter_type)
            if not specs_field:
                continue
            key = (
                rec.make,
                rec.model,
                rec.year_from,
                rec.year_to,
                rec.engine_code,
            )
            for oem in rec.oem_parts:
                if oem:
                    groups[key][specs_field].add(oem)

        results: list[dict] = []
        for (make, model, y_from, y_to, eng), fields in groups.items():
            entry: dict = {
                "make": make,
                "model": model,
                "year_from": y_from,
                "year_to": y_to,
                "engine_code": eng,
            }
            for field_name, parts in fields.items():
                # Join with comma, sorted for determinism
                entry[field_name] = ", ".join(sorted(parts))
            results.append(entry)
        return results

    # -- internals ---------------------------------------------------------

    def _detect_header(self, ws) -> tuple[dict[int, str], int]:
        """Scan the first N rows for a header row.

        Returns ``(col_map, header_row_number)`` where *col_map* maps
        1-based column indices to canonical field names.

        Raises ``ValueError`` if no qualifying header row is found.
        """
        for row_idx, row in enumerate(ws.iter_rows(max_row=self.max_header_scan_rows), start=1):
            col_map: dict[int, str] = {}
            for col_idx, cell in enumerate(row, start=1):
                raw = _cell_str(cell.value).lower()
                if raw in COLUMN_ALIASES:
                    col_map[col_idx] = COLUMN_ALIASES[raw]

            has_required = "make" in col_map.values() or "filtron_part" in col_map.values()
            if len(col_map) >= 3 and has_required:
                logger.debug("Header detected at row %d: %s", row_idx, col_map)
                return col_map, row_idx

        raise ValueError(
            f"No header row found in first {self.max_header_scan_rows} rows "
            f"of '{self.path.name}'. Expected >= 3 recognized columns "
            f"including 'make' or 'filtron_part'."
        )

    def _parse_rows(
        self,
        ws,
        col_map: dict[int, str],
        header_row: int,
    ) -> list[FiltronRecord]:
        """Parse data rows below the header into FiltronRecord objects."""
        records: list[FiltronRecord] = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
            raw: dict[str, str] = {}
            for col_idx, canonical in col_map.items():
                cell = row[col_idx - 1] if col_idx <= len(row) else None
                raw[canonical] = _cell_str(cell.value if cell else None)

            # Skip fully empty rows
            if not any(raw.values()):
                continue

            rec = FiltronRecord()
            rec.make = raw.get("make", "")
            rec.model = raw.get("model", "")
            rec.engine_code = raw.get("engine_code", "")
            rec.filtron_part = raw.get("filtron_part", "")
            rec.mann_part = raw.get("mann_part", "")

            # Year handling: prefer separate from/to, fall back to range
            if raw.get("year_from"):
                rec.year_from = _safe_int(raw["year_from"])
                rec.year_to = _safe_int(raw.get("year_to", "")) or OPEN_ENDED_YEAR
            elif raw.get("year_range"):
                rec.year_from, rec.year_to = parse_year_range(raw["year_range"])

            # OEM parts: may be multi-valued
            rec.oem_parts = split_multi_part(raw.get("oem_part", ""))

            # Filter type: explicit > inferred from part prefix
            rec.filter_type = resolve_filter_type(
                raw.get("filter_type", ""), rec.filtron_part
            )

            records.append(rec)

        return records


def _safe_int(value: str) -> int:
    """Parse an integer, returning 0 on failure."""
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return 0
